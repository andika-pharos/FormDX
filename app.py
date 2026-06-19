#!/usr/bin/env python3
"""
FormDX v1.0 - Clean Edition
Sistem Manajemen R&D Formulasi dengan Stock Card & Autocomplete
"""

from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, g, make_response
import sqlite3
from datetime import datetime, date
from functools import wraps
import io
import csv
import json

app = Flask(__name__)
app.secret_key = 'formdx-v1-clean-2026'
app.config['DATABASE'] = 'formdx.db'

def get_db():
    """Mendapatkan koneksi database SQLite dengan pengelolaan yang lebih baik.
    - Di dalam request: menggunakan flask.g (connection pooling per request)
    - Di luar request (startup): membuat koneksi langsung
    """
    # Cek apakah sedang di dalam application context
    from flask import has_app_context

    if has_app_context():
        # Jika belum ada koneksi di g, atau koneksi sebelumnya sudah ditutup
        if 'db' not in g or g.db is None:
            conn = sqlite3.connect(app.config['DATABASE'], timeout=30)
            conn.row_factory = sqlite3.Row
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("PRAGMA busy_timeout = 30000")
            g.db = conn
            return g.db

        # Cek apakah koneksi masih bisa digunakan (jika sebelumnya di-close manual)
        try:
            g.db.execute("SELECT 1")
            return g.db
        except sqlite3.ProgrammingError:
            # Koneksi sudah closed → buat koneksi baru
            conn = sqlite3.connect(app.config['DATABASE'], timeout=30)
            conn.row_factory = sqlite3.Row
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("PRAGMA busy_timeout = 30000")
            g.db = conn
            return g.db
    else:
        # Diluar application context (misalnya saat init_db di startup)
        conn = sqlite3.connect(app.config['DATABASE'], timeout=30)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA busy_timeout = 30000")
        return conn


@app.teardown_appcontext
def close_db(e=None):
    """Menutup koneksi database di akhir setiap request."""
    db = g.pop('db', None)
    if db is not None:
        db.close()

def init_db():
    conn = get_db()
    c = conn.cursor()

    # Tabel Bahan Baku
    c.execute('''
        CREATE TABLE IF NOT EXISTS raw_materials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL, name TEXT NOT NULL, description TEXT,
            supplier TEXT, manufacturer TEXT, unit TEXT DEFAULT 'g',
            current_stock REAL DEFAULT 0, min_stock REAL DEFAULT 0,
            price_per_unit REAL DEFAULT 0, expiry_date TEXT,
            lpb_number TEXT, lsa_number TEXT, location TEXT DEFAULT 'R&D Coolroom',
            created_at TEXT, updated_at TEXT
        )
    ''')

    # Tabel Trial
    c.execute('''
        CREATE TABLE IF NOT EXISTS trials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            trial_code TEXT UNIQUE NOT NULL, master_batch_id TEXT,
            trial_date TEXT NOT NULL, project_name TEXT NOT NULL,
            objective TEXT, formulator TEXT,
            batch_size_value REAL DEFAULT 0, batch_size_unit TEXT DEFAULT 'tablet',
            scheme INTEGER DEFAULT 1, tahap TEXT DEFAULT 'F1',
            procedure TEXT, observations TEXT, results TEXT, conclusion TEXT,
            status TEXT DEFAULT 'Draft', created_at TEXT, updated_at TEXT
        )
    ''')

    # Tabel Komposisi
    c.execute('''
        CREATE TABLE IF NOT EXISTS trial_ingredients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            trial_id INTEGER NOT NULL, raw_material_id INTEGER NOT NULL,
            quantity REAL NOT NULL, lot_number TEXT, notes TEXT,
            qty_per_small_unit REAL, small_unit TEXT, phase TEXT DEFAULT 'Fase Utama',
            FOREIGN KEY (trial_id) REFERENCES trials (id) ON DELETE CASCADE
        )
    ''')

    # Tabel Stock Movements (Kartu Stok)
    c.execute('''
        CREATE TABLE IF NOT EXISTS stock_movements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            raw_material_id INTEGER NOT NULL,
            movement_type TEXT NOT NULL,           -- 'IN' atau 'OUT'
            quantity REAL NOT NULL,
            unit TEXT NOT NULL,
            reference_type TEXT,                   -- 'TRIAL', 'INBOUND', 'ADJUSTMENT'
            reference_id INTEGER,
            reference_code TEXT,                   -- master_batch_id atau trial_code
            project_name TEXT,
            notes TEXT,
            created_at TEXT,
            lpb_number TEXT,
            lsa_number TEXT,
            expiry_date TEXT,
            FOREIGN KEY (raw_material_id) REFERENCES raw_materials(id)
        )
    ''')

    # Tambah kolom baru jika tabel sudah ada (untuk backward compatibility)
    try:
        c.execute("ALTER TABLE stock_movements ADD COLUMN lpb_number TEXT")
    except:
        pass
    try:
        c.execute("ALTER TABLE stock_movements ADD COLUMN lsa_number TEXT")
    except:
        pass
    try:
        c.execute("ALTER TABLE stock_movements ADD COLUMN expiry_date TEXT")
    except:
        pass

    # Tabel Riwayat Perubahan Trial (Audit Log)
    c.execute('''
        CREATE TABLE IF NOT EXISTS trial_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            trial_id INTEGER NOT NULL,
            action TEXT NOT NULL,           -- CREATE, UPDATE, DELETE
            changed_by TEXT,
            changed_at TEXT,
            old_data TEXT,
            new_data TEXT,
            notes TEXT,
            FOREIGN KEY (trial_id) REFERENCES trials(id) ON DELETE CASCADE
        )
    ''')

    conn.commit()
    # Jangan close manual di sini. Biarkan @app.teardown_appcontext yang menutup koneksi
    # setelah with app.app_context(): selesai.

def add_sample_data():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM raw_materials")
    if c.fetchone()[0] > 0:
        # Jangan close manual. Biarkan teardown yang handle.
        return
    now = datetime.now().isoformat()
    # Sample data bahan baku untuk tablet (kode 6 digit random)
    # Tambahkan No LPB (B + 6 digit) dan No LSA (O + 6 digit) untuk setiap bahan
    samples = [
        ('482917', 'Microcrystalline Cellulose', 'Filler / Diluent', 'FMC Biopolymer', 'FMC Biopolymer', 'kg', 125.0, 30.0, 185000, '2028-03-15', 'B482917', 'O139204'),
        ('139204', 'Lactose Monohydrate', 'Diluent', 'DFE Pharma', 'DFE Pharma', 'kg', 98.5, 25.0, 125000, '2027-11-20', 'B139204', 'O756381'),
        ('756381', 'Magnesium Stearate', 'Lubricant', 'Peter Greven', 'Peter Greven', 'kg', 12.8, 3.0, 285000, '2027-08-10', 'B756381', 'O294756'),
        ('294756', 'Croscarmellose Sodium', 'Disintegrant', 'FMC Biopolymer', 'FMC Biopolymer', 'kg', 8.5, 2.0, 320000, '2028-01-05', 'B294756', 'O617492'),
        ('617492', 'Povidone K30', 'Binder', 'BASF', 'BASF', 'kg', 15.2, 4.0, 245000, '2027-09-18', 'B617492', 'O831569'),
        ('831569', 'Colloidal Silicon Dioxide', 'Glidant', 'Evonik', 'Evonik', 'kg', 5.8, 1.5, 385000, '2028-02-28', 'B831569', 'O475283'),
        ('475283', 'Talc', 'Glidant / Lubricant', 'Imerys', 'Imerys', 'kg', 22.0, 5.0, 95000, '2027-12-12', 'B475283', 'O928174'),
        ('928174', 'Sodium Starch Glycolate', 'Disintegrant', 'Roquette', 'Roquette', 'kg', 9.3, 2.5, 275000, '2028-04-08', 'B928174', 'O361849'),
        ('361849', 'Mannitol', 'Diluent / Sweetener', 'Roquette', 'Roquette', 'kg', 45.0, 10.0, 165000, '2027-10-25', 'B361849', 'O584732'),
        ('584732', 'Stearic Acid', 'Lubricant', 'Peter Greven', 'Peter Greven', 'kg', 7.5, 2.0, 195000, '2027-07-14', 'B584732', 'O482917'),
    ]
    for s in samples:
        code, name, description, supplier, manufacturer, unit, current_stock, min_stock, price, expiry, lpb1, lsa1 = s
        
        # Insert material with first lot as main reference
        c.execute('''INSERT INTO raw_materials 
            (code, name, description, supplier, manufacturer, unit, current_stock, min_stock, price_per_unit, expiry_date, lpb_number, lsa_number, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', 
            (code, name, description, supplier, manufacturer, unit, current_stock, min_stock, price, expiry, lpb1, lsa1, now, now))
        
        material_id = c.lastrowid
        
        # Create 2 inbound movements with different LPB/LSA and split stock (for demo multiple lots)
        stock1 = round(current_stock * 0.6, 2)   # 60% in first lot
        stock2 = round(current_stock * 0.4, 2)   # 40% in second lot
        
        lpb2 = 'B' + ''.join([str(__import__('random').randint(0,9)) for _ in range(6)])
        lsa2 = 'O' + ''.join([str(__import__('random').randint(0,9)) for _ in range(6)])
        
        # ED untuk masing-masing lot (berbeda)
        ed1 = expiry
        ed2 = (datetime.fromisoformat(expiry) - __import__('datetime').timedelta(days=180)).date().isoformat() if expiry else None

        # First inbound
        c.execute('''
            INSERT INTO stock_movements 
            (raw_material_id, movement_type, quantity, unit, reference_type, reference_code, lpb_number, lsa_number, expiry_date, notes, created_at)
            VALUES (?, 'IN', ?, ?, 'INBOUND', ?, ?, ?, ?, ?, ?)
        ''', (material_id, stock1, unit, lpb1, lpb1, lsa1, ed1, f"Initial stock - Lot 1", now))
        
        # Second inbound (different lot)
        c.execute('''
            INSERT INTO stock_movements 
            (raw_material_id, movement_type, quantity, unit, reference_type, reference_code, lpb_number, lsa_number, expiry_date, notes, created_at)
            VALUES (?, 'IN', ?, ?, 'INBOUND', ?, ?, ?, ?, ?, ?)
        ''', (material_id, stock2, unit, lpb2, lpb2, lsa2, ed2, f"Initial stock - Lot 2", now))
    
    conn.commit()
    print("Sample data added with multiple lots per material (2 LPB & 2 LSA each).")
    # Koneksi akan ditutup otomatis oleh teardown_appcontext

def log_trial_history(trial_id, action, changed_by=None, old_data=None, new_data=None, notes=None, conn=None):
    """Mencatat riwayat perubahan trial"""
    close_conn = False
    if conn is None:
        conn = get_db()
        close_conn = True

    c = conn.cursor()
    now = datetime.now().isoformat()

    old_json = json.dumps(old_data, default=str) if old_data else None
    new_json = json.dumps(new_data, default=str) if new_data else None

    c.execute('''
        INSERT INTO trial_history (trial_id, action, changed_by, changed_at, old_data, new_data, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (trial_id, action, changed_by, now, old_json, new_json, notes))

    if close_conn:
        conn.commit()
        conn.close()


def log_stock_movement(raw_material_id, movement_type, quantity, unit, reference_type=None, reference_id=None, reference_code=None, project_name=None, notes=None, lpb_number=None, lsa_number=None, conn=None):
    """Mencatat pergerakan stok ke tabel stock_movements.
    Jika conn diberikan, akan menggunakan koneksi tersebut (untuk transaksi besar).
    Jika tidak, akan membuat koneksi baru.
    """
    close_conn = False
    if conn is None:
        conn = get_db()
        close_conn = True

    c = conn.cursor()
    now = datetime.now().isoformat()
    c.execute('''
        INSERT INTO stock_movements 
        (raw_material_id, movement_type, quantity, unit, reference_type, reference_id, reference_code, project_name, lpb_number, lsa_number, notes, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (raw_material_id, movement_type, quantity, unit, reference_type, reference_id, reference_code, project_name, lpb_number, lsa_number, notes, now))
    
    if close_conn:
        conn.commit()
        conn.close()
    # Jika conn diberikan dari luar, jangan commit/close di sini (biarkan pemanggil yang atur)

def convert_quantity(quantity, from_unit, to_unit):
    if quantity is None: return 0.0
    try: q = float(quantity)
    except: return 0.0
    from_u = str(from_unit).lower().strip()
    to_u = str(to_unit).lower().strip()
    mass = {'mg': 0.001, 'g': 1.0, 'kg': 1000.0}
    vol = {'ml': 1.0, 'l': 1000.0, 'mL': 1.0, 'L': 1000.0}
    if from_u in mass and to_u in mass:
        return round(q * (mass[from_u] / mass[to_u]), 4)
    elif from_u in vol and to_u in vol:
        return round(q * (vol[from_u] / vol[to_u]), 4)
    return round(q, 4)

# ==================== AUTH ====================
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            flash('Silakan login terlebih dahulu.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

USERS = {
    'formulator': {'password': 'rd2026', 'role': 'scientist', 'name': 'Formulation Scientist'},
    'admin': {'password': 'admin123', 'role': 'admin', 'name': 'Admin R&D'}
}

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if username in USERS and USERS[username]['password'] == password:
            session['user'] = username
            session['role'] = USERS[username]['role']
            session['name'] = USERS[username]['name']
            flash(f'Selamat datang, {session["name"]}!', 'success')
            return redirect(url_for('dashboard'))
        flash('Username atau password salah.', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Anda telah logout.', 'info')
    return redirect(url_for('login'))

# ==================== ROUTES ====================

@app.route('/')
def index():
    return redirect(url_for('dashboard') if 'user' in session else url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM raw_materials"); total_materials = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM trials"); total_trials = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM raw_materials WHERE current_stock < min_stock"); low_stock_count = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM trials WHERE status = 'Draft'"); draft_trials = c.fetchone()[0]
    c.execute("SELECT id, code, name, current_stock, min_stock, unit FROM raw_materials WHERE current_stock < min_stock ORDER BY (current_stock - min_stock) ASC LIMIT 5")
    low_stock_materials = c.fetchall()
    from datetime import datetime, timedelta

    today = datetime.now().date()
    three_months_later = (today + timedelta(days=90)).isoformat()

    # Counter bahan mendekati ED (< 3 bulan)
    c.execute("""
        SELECT COUNT(*) FROM raw_materials 
        WHERE expiry_date IS NOT NULL 
        AND expiry_date <= ?
    """, (three_months_later,))
    expiring_soon_count = c.fetchone()[0]

    # List bahan mendekati ED
    c.execute("""
        SELECT code, name, expiry_date, current_stock, unit 
        FROM raw_materials 
        WHERE expiry_date IS NOT NULL 
        AND expiry_date <= ?
        ORDER BY expiry_date ASC 
        LIMIT 8
    """, (three_months_later,))
    expiring_materials = c.fetchall()

    # Produk + Jumlah Formula/Trial
    c.execute("""
        SELECT project_name, COUNT(*) as formula_count 
        FROM trials 
        WHERE project_name IS NOT NULL 
        GROUP BY project_name 
        ORDER BY formula_count DESC 
        LIMIT 8
    """)
    product_formulas = c.fetchall()

    conn.close()
    from datetime import datetime
    current_date = datetime.now().strftime("%d %B %Y")

    return render_template('dashboard.html', total_materials=total_materials, total_trials=total_trials,
                           low_stock_count=low_stock_count, draft_trials=draft_trials,
                           low_stock_materials=low_stock_materials,
                           expiring_materials=expiring_materials,
                           expiring_soon_count=expiring_soon_count,
                           product_formulas=product_formulas,
                           current_date=current_date)

# ==================== RAW MATERIALS + STOCK CARD ====================

@app.route('/raw_materials')
@login_required
def raw_materials():
    conn = get_db()
    c = conn.cursor()
    search = request.args.get('search', '').strip()
    low_stock = request.args.get('low_stock', '')

    query = """
        SELECT rm.*, 
               (SELECT COUNT(*) FROM stock_movements sm 
                WHERE sm.raw_material_id = rm.id AND sm.movement_type = 'IN') as lot_count
        FROM raw_materials rm
    """
    params = []
    conditions = []

    if search:
        conditions.append("(rm.code LIKE ? OR rm.name LIKE ? OR rm.supplier LIKE ?)")
        params.extend([f'%{search}%'] * 3)

    if low_stock == '1':
        conditions.append("rm.current_stock < rm.min_stock")

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += " ORDER BY rm.name ASC"
    c.execute(query, params)
    materials = c.fetchall()

    # Ambil semua master data untuk dropdown di modal Tambah Bahan + Stok
    c.execute("SELECT id, code, name, description, manufacturer, supplier FROM raw_materials ORDER BY code ASC")
    all_materials = c.fetchall()

    conn.close()

    return render_template('raw_materials.html', materials=materials, all_materials=all_materials, search=search, filter_low=low_stock)


# ==================== RAW MATERIALS CRUD (dibutuhkan oleh template) ====================

@app.route('/raw_materials/add', methods=['POST'])
@login_required
def add_raw_material():
    code = request.form.get('code', '').strip().upper()
    name = request.form.get('name', '').strip()
    if not code or not name:
        flash('Kode dan Nama bahan wajib diisi!', 'danger')
        return redirect(url_for('raw_materials'))

    now = datetime.now().isoformat()
    conn = get_db()
    c = conn.cursor()

    try:
        # Cek apakah kode sudah ada
        c.execute("SELECT id, current_stock FROM raw_materials WHERE code = ?", (code,))
        existing = c.fetchone()

        if existing:
            # Jika sudah ada → Update stok + catat movement Inbound (dengan konversi satuan)
            material_id = existing['id']
            old_stock = existing['current_stock']

            # Ambil satuan master bahan
            c.execute("SELECT unit FROM raw_materials WHERE id = ?", (material_id,))
            master_unit = c.fetchone()['unit']

            input_qty = float(request.form.get('current_stock', 0) or 0)
            input_unit = request.form.get('unit', 'g').strip().lower()

            # Konversi ke satuan master
            converted_qty = convert_quantity(input_qty, input_unit, master_unit)

            c.execute('''
                UPDATE raw_materials 
                SET current_stock = current_stock + ?, updated_at = ?
                WHERE id = ?
            ''', (converted_qty, now, material_id))

            # Catat movement dengan satuan master
            lpb = request.form.get('lpb_number', '').strip().upper()
            lsa = request.form.get('lsa_number', '').strip().upper()
            notes = f"Tambah stok via form. Old: {old_stock}, Added: {converted_qty} {master_unit}"

            c.execute('''
                INSERT INTO stock_movements 
                (raw_material_id, movement_type, quantity, unit, reference_type, reference_code, 
                 lpb_number, lsa_number, notes, created_at)
                VALUES (?, 'IN', ?, ?, 'INBOUND', ?, ?, ?, ?, ?)
            ''', (material_id, converted_qty, master_unit,
                  lpb or code, lpb, lsa, notes, now))

            conn.commit()
            flash(f'Stok bahan "{name}" berhasil ditambahkan sebanyak {converted_qty} {master_unit}.', 'success')

        else:
            # Jika belum ada → Insert bahan baru
            c.execute('''
                INSERT INTO raw_materials 
                (code, name, description, supplier, manufacturer, unit, current_stock, min_stock, 
                 price_per_unit, expiry_date, lpb_number, lsa_number, location, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                code,
                name.upper(),
                request.form.get('description', '').strip().upper(),
                request.form.get('supplier', '').strip().upper(),
                request.form.get('manufacturer', '').strip().upper(),
                request.form.get('unit', 'g').strip().upper(),
                float(request.form.get('current_stock', 0) or 0),
                float(request.form.get('min_stock', 0) or 0),
                float(request.form.get('price_per_unit', 0) or 0),
                request.form.get('expiry_date', ''),
                request.form.get('lpb_number', '').strip().upper(),
                request.form.get('lsa_number', '').strip().upper(),
                request.form.get('location', 'R&D Coolroom').strip().upper(),
                now, now
            ))
            conn.commit()
            flash(f'Bahan baku baru "{name}" berhasil ditambahkan.', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Gagal menambahkan: {str(e)}', 'danger')
    finally:
        conn.close()

    return redirect(url_for('raw_materials'))


@app.route('/raw_materials/edit/<int:id>', methods=['POST'])
@login_required
def edit_raw_material(id):
    now = datetime.now().isoformat()
    conn = get_db()
    c = conn.cursor()

    # Ambil stok lama
    c.execute("SELECT current_stock, name FROM raw_materials WHERE id = ?", (id,))
    old_data = c.fetchone()
    old_stock = old_data['current_stock'] if old_data else 0
    material_name = old_data['name'] if old_data else ''

    new_stock = float(request.form.get('current_stock', 0) or 0)
    difference = new_stock - old_stock

    c.execute('''
        UPDATE raw_materials SET
            name=?, description=?, supplier=?, manufacturer=?, unit=?,
            current_stock=?, min_stock=?, price_per_unit=?, expiry_date=?,
            lpb_number=?, lsa_number=?, location=?, updated_at=?
        WHERE id=?
    ''', (
        request.form.get('name', '').strip().upper(),
        request.form.get('description', '').strip().upper(),
        request.form.get('supplier', '').strip().upper(),
        request.form.get('manufacturer', '').strip().upper(),
        request.form.get('unit', 'g').strip().upper(),
        new_stock,
        float(request.form.get('min_stock', 0) or 0),
        float(request.form.get('price_per_unit', 0) or 0),
        request.form.get('expiry_date', ''),
        request.form.get('lpb_number', '').strip().upper(),
        request.form.get('lsa_number', '').strip().upper(),
        request.form.get('location', 'R&D Coolroom').strip().upper(),
        now, id
    ))

    # Catat adjustment jika ada perubahan stok
    if difference != 0:
        adjustment_type = 'IN' if difference > 0 else 'OUT'
        log_stock_movement(
            raw_material_id=id,
            movement_type=adjustment_type,
            quantity=abs(difference),
            unit=request.form.get('unit', 'g').strip(),
            reference_type='ADJUSTMENT',
            reference_id=id,
            reference_code='STOCK_EDIT',
            project_name=material_name,
            notes=f"Stock adjustment via edit form. Old: {old_stock}, New: {new_stock}",
            conn=conn
        )

    conn.commit()
    conn.close()
    flash('Data bahan baku berhasil diperbarui dan perubahan stok tercatat di Kartu Stok.', 'success')
    return redirect(url_for('raw_materials'))


@app.route('/raw_materials/delete/<int:id>', methods=['POST'])
@login_required
def delete_raw_material(id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM raw_materials WHERE id=?", (id,))
    conn.commit()
    conn.close()
    flash('Bahan baku berhasil dihapus.', 'warning')
    return redirect(url_for('raw_materials'))


@app.route('/stock_card/<int:material_id>')
@login_required
def stock_card(material_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM raw_materials WHERE id = ?", (material_id,))
    material = c.fetchone()
    if not material:
        flash('Bahan baku tidak ditemukan.', 'danger')
        return redirect(url_for('raw_materials'))
    
    c.execute("""
        SELECT * FROM stock_movements 
        WHERE raw_material_id = ? 
        ORDER BY created_at DESC
    """, (material_id,))
    movements = c.fetchall()
    conn.close()
    return render_template('stock_card.html', material=material, movements=movements)

@app.route('/stock_card/export/<int:material_id>')
@login_required
def export_stock_card(material_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM raw_materials WHERE id = ?", (material_id,))
    material = c.fetchone()
    c.execute("""
        SELECT created_at, movement_type, quantity, unit, reference_type, reference_code, 
               project_name, notes, lpb_number, lsa_number, expiry_date
        FROM stock_movements 
        WHERE raw_material_id = ? ORDER BY created_at ASC
    """, (material_id,))
    movements = c.fetchall()
    conn.close()

    # Hitung stok awal
    total_in = sum(m['quantity'] for m in movements if m['movement_type'] == 'IN')
    total_out = sum(m['quantity'] for m in movements if m['movement_type'] == 'OUT')
    opening_stock = material['current_stock'] - total_in + total_out

    wb = Workbook()
    ws = wb.active
    ws.title = "Kartu Stok"

    # Styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    title_font = Font(bold=True, size=16, color="0d3b66")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="0d3b66", end_color="0d3b66", fill_type="solid")
    light_fill = PatternFill(start_color="e8f4f8", end_color="e8f4f8", fill_type="solid")

    # === HEADER ===
    ws.merge_cells('A1:J1')
    ws['A1'] = "FormDX v1.0 - KARTU STOK BAHAN BAKU"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:J2')
    ws['A2'] = "Dokumen ini dibuat untuk keperluan R&D dan pelaporan internal"
    ws['A2'].alignment = center_align

    # Info Bahan
    ws['A4'] = "Kode Bahan"
    ws['B4'] = material['code']
    ws['D4'] = "Nama Bahan"
    ws['E4'] = material['name']
    ws['A5'] = "Satuan Dasar"
    ws['B5'] = material['unit']
    ws['D5'] = "Tanggal Kadaluarsa (ED)"
    ws['E5'] = material['expiry_date'] or '-'

    for row in range(4, 6):
        for col in ['A', 'D']:
            ws[f'{col}{row}'].font = bold_font
            ws[f'{col}{row}'].fill = light_fill

    # Header tabel pergerakan
    headers = ["No", "Tanggal", "Waktu", "Tipe", "Jumlah Masuk", "Jumlah Keluar", 
               "Satuan", "Saldo Akhir", "No. LPB", "No. LSA", "Keterangan"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Set column widths (more balanced)
    column_widths = {
        'A': 6,   # No
        'B': 12,  # Tanggal
        'C': 8,   # Waktu
        'D': 10,  # Tipe
        'E': 12,  # Jumlah Masuk
        'F': 12,  # Jumlah Keluar
        'G': 8,   # Satuan
        'H': 12,  # Saldo Akhir
        'I': 14,  # No. LPB
        'J': 14,  # No. LSA
        'K': 35,  # Keterangan
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Data pergerakan
    balance = opening_stock
    for idx, m in enumerate(movements, 1):
        qty_in = ''
        qty_out = ''
        if m['movement_type'] == 'IN':
            qty_in = m['quantity']
            balance += m['quantity']
        else:
            qty_out = m['quantity']
            balance -= m['quantity']

        created_at = m['created_at'] or ''
        date_part = created_at[:10] if len(created_at) >= 10 else ''
        time_part = created_at[11:16] if len(created_at) >= 16 else ''

        row_data = [
            idx,
            date_part,
            time_part,
            m['movement_type'],
            qty_in,
            qty_out,
            m['unit'],
            round(balance, 2),
            m['lpb_number'] or m['reference_code'] or '',
            m['lsa_number'] or '',
            m['notes'] or ''
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=7 + idx, column=col, value=value)
            cell.alignment = center_align
            cell.border = thin_border

    # Auto column width
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.column_dimensions['J'].width = 35

    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename=StockCard_{material['code']}.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


@app.route('/stock_card/<int:material_id>/export_pdf')
@login_required
def export_stock_card_pdf(material_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import cm
    import io
    from flask import make_response
    from datetime import datetime

    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM raw_materials WHERE id = ?", (material_id,))
    material = c.fetchone()

    if not material:
        conn.close()
        flash('Bahan baku tidak ditemukan.', 'danger')
        return redirect(url_for('raw_materials'))

    c.execute("""
        SELECT created_at, movement_type, quantity, unit, reference_type, reference_code, 
               project_name, notes, lpb_number, lsa_number, expiry_date
        FROM stock_movements 
        WHERE raw_material_id = ? ORDER BY created_at ASC
    """, (material_id,))
    movements = c.fetchall()
    conn.close()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16,
                                  textColor=colors.HexColor('#0d3b66'), alignment=1, spaceAfter=4)

    elements = []

    # Header
    elements.append(Paragraph("KARTU STOK BAHAN BAKU", title_style))
    elements.append(Spacer(1, 0.4*cm))

    # Material Info
    info_data = [
        ["Kode Bahan", material['code'], "Nama Bahan", material['name']],
        ["Satuan Dasar", material['unit'], "Tanggal Kadaluarsa (ED)", material['expiry_date'] or '-'],
        ["Lokasi Penyimpanan", material['location'] or '-', "Supplier", material['supplier'] or '-']
    ]
    info_table = Table(info_data, colWidths=[3.8*cm, 5.2*cm, 3.8*cm, 5.2*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f4f8')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#e8f4f8')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.4*cm))

    # Movements
    elements.append(Paragraph("<b>RIWAYAT PERGERAKAN STOK</b>", styles['Heading2']))

    if movements:
        # Hitung stok awal sebelum semua pergerakan
        total_in = sum(m['quantity'] for m in movements if m['movement_type'] == 'IN')
        total_out = sum(m['quantity'] for m in movements if m['movement_type'] == 'OUT')
        opening_stock = material['current_stock'] - total_in + total_out

        movement_data = [["No", "Tanggal", "Tipe", "Masuk", "Keluar", "Satuan", "Stok Akhir", "No. LPB", "No. LSA", "ED", "Keterangan"]]
        balance = opening_stock

        for idx, m in enumerate(movements, 1):
            date_str = m['created_at'][:10] if m['created_at'] else '-'
            qty_in = ""
            qty_out = ""

            if m['movement_type'] == 'IN':
                qty_in = f"{m['quantity']:.2f}"
                balance += m['quantity']
            else:
                qty_out = f"{m['quantity']:.2f}"
                balance -= m['quantity']

            lpb = m['lpb_number'] or m['reference_code'] or '-'
            lsa = m['lsa_number'] or '-'
            ed = m['expiry_date'] or '-'
            keterangan = m['notes'] or m['project_name'] or '-'

            movement_data.append([
                str(idx),
                date_str,
                m['movement_type'],
                qty_in,
                qty_out,
                m['unit'],
                f"{balance:.2f}",  # This is now "Stok Akhir"
                lpb,
                lsa,
                ed,
                keterangan[:45]
            ])

        move_table = Table(movement_data, colWidths=[0.55*cm, 1.4*cm, 0.95*cm, 1.2*cm, 1.2*cm, 1.0*cm, 1.4*cm, 1.9*cm, 1.9*cm, 1.8*cm, 5.8*cm])
        move_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d3b66')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (7, 1), (7, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#999999')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(move_table)
    else:
        elements.append(Paragraph("<i>Belum ada riwayat pergerakan stok untuk bahan ini.</i>", styles['Normal']))

    doc.build(elements)

    buffer.seek(0)
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    filename = f"StockCard_{material['code']}.pdf"
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response


# ==================== API untuk Autocomplete ====================

@app.route('/api/search_materials')
@login_required
def search_materials():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT id, code, name, unit, current_stock, lsa_number 
        FROM raw_materials 
        WHERE code LIKE ? OR name LIKE ? 
        ORDER BY name LIMIT 15
    """, (f'%{q}%', f'%{q}%'))
    results = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(results)


@app.route('/api/material_lots/<int:material_id>')
@login_required
def get_material_lots(material_id):
    """Mengambil daftar lot (LSA) aktif untuk suatu bahan baku"""
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT DISTINCT lsa_number, lpb_number, expiry_date
        FROM stock_movements 
        WHERE raw_material_id = ? 
          AND lsa_number IS NOT NULL 
          AND movement_type = 'IN'
        ORDER BY created_at DESC
    """, (material_id,))
    lots = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(lots)


# ==================== TRIALS ====================

def generate_trial_code():
    now = datetime.now()
    year_month = now.strftime('%Y%m')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM trials WHERE trial_code LIKE ?", (f'TR-{year_month}-%',))
    count = c.fetchone()[0] + 1
    # Jangan conn.close() di sini! Biarkan teardown_appcontext yang menutup di akhir request.
    return f"TR-{year_month}-{count:03d}"

def generate_master_batch_id(trial_date_str):
    try:
        trial_date = datetime.fromisoformat(trial_date_str)
    except:
        trial_date = datetime.now()
    year_suffix = trial_date.strftime('%y')
    month_letter = chr(ord('A') + trial_date.month - 1)
    conn = get_db()
    c = conn.cursor()
    pattern = f'F____{month_letter}{year_suffix}A'
    c.execute("SELECT COUNT(*) FROM trials WHERE master_batch_id LIKE ?", (pattern,))
    count = c.fetchone()[0] + 1
    # Jangan conn.close() di sini! Biarkan teardown_appcontext yang menutup di akhir request.
    return f"F{count:04d}{month_letter}{year_suffix}A"

@app.route('/trials')
@login_required
def trials_list():
    conn = get_db()
    c = conn.cursor()
    search = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    query = """
        SELECT t.*, (SELECT COUNT(*) FROM trial_ingredients ti WHERE ti.trial_id = t.id) as ingredient_count
        FROM trials t
    """
    params = []
    conditions = []
    if search:
        conditions.append("(t.trial_code LIKE ? OR t.project_name LIKE ?)")
        params.extend([f'%{search}%']*2)
    if status_filter:
        conditions.append("t.status = ?")
        params.append(status_filter)
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY t.project_name ASC, t.trial_date DESC"
    c.execute(query, params)
    trials = c.fetchall()
    conn.close()

    from itertools import groupby
    from operator import itemgetter
    grouped_trials = []
    for key, group in groupby(trials, key=itemgetter('project_name')):
        grouped_trials.append({'project_name': key, 'trials': list(group)})
    return render_template('trials_list.html', grouped_trials=grouped_trials, search=search, status_filter=status_filter)

@app.route('/trials/create', methods=['GET', 'POST'])
@login_required
def create_trial():
    if request.method == 'POST':
        trial_date = request.form.get('trial_date')
        project_name = request.form.get('project_name', '').strip().upper()
        if not project_name or not trial_date:
            flash('Nama Proyek dan Tanggal Trial wajib diisi!', 'danger')
            return redirect(url_for('create_trial'))

        batch_size_value = float(request.form.get('batch_size_value', 0) or 0)
        batch_size_unit = request.form.get('batch_size_unit', 'tablet')
        update_stock = request.form.get('update_stock') == 'on'

        trial_code = generate_trial_code()
        master_batch_id = generate_master_batch_id(trial_date)
        now = datetime.now().isoformat()

        conn = get_db()
        c = conn.cursor()
        try:
            c.execute('''
                INSERT INTO trials (trial_code, master_batch_id, trial_date, project_name, objective, formulator,
                    batch_size_value, batch_size_unit, scheme, tahap, procedure, status, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (trial_code, master_batch_id, trial_date, project_name,
                  request.form.get('objective', '').strip(),
                  request.form.get('formulator', session.get('name', '')).strip().upper(),
                  batch_size_value, batch_size_unit,
                  int(request.form.get('scheme', 1)),
                  request.form.get('tahap', 'F1'),
                  request.form.get('procedure', '').strip(),
                  request.form.get('status', 'Draft'), now, now))
            trial_id = c.lastrowid

            raw_ids = request.form.getlist('raw_material_id[]')
            qty_per_smalls = request.form.getlist('qty_per_small_unit[]')
            small_units = request.form.getlist('small_unit[]')
            phases = request.form.getlist('phase[]')
            lot_numbers = request.form.getlist('lot_number[]')

            # === TWO-PASS STOCK CHECK (lebih aman) ===
            ingredients_data = []
            insufficient = []

            for i, raw_id in enumerate(raw_ids):
                if not raw_id or not qty_per_smalls[i]: continue
                qty_per_small = float(qty_per_smalls[i])
                small_unit = small_units[i] if i < len(small_units) else 'mg'
                phase = phases[i] if i < len(phases) else 'Fase Utama'
                lot_number = lot_numbers[i] if i < len(lot_numbers) else ''
                total_needed = round(qty_per_small * batch_size_value, 4)

                ingredients_data.append({
                    'raw_id': int(raw_id),
                    'total_needed': total_needed,
                    'small_unit': small_unit,
                    'phase': phase,
                    'lot_number': lot_number,
                    'qty_per_small': qty_per_small
                })

                if update_stock:
                    c.execute("SELECT unit, current_stock, name FROM raw_materials WHERE id = ?", (int(raw_id),))
                    mat = c.fetchone()
                    if mat:
                        converted = convert_quantity(total_needed, small_unit, mat['unit'])
                        if mat['current_stock'] < converted:
                            insufficient.append(f"{mat['name']} (butuh {converted:.2f} {mat['unit']})")

            if insufficient:
                conn.rollback()
                flash("Stok tidak mencukupi:<br>" + "<br>".join(insufficient), 'danger')
                return redirect(url_for('create_trial'))

            # === Jika semua stok cukup, baru proses ===
            for ing in ingredients_data:
                c.execute('''INSERT INTO trial_ingredients 
                    (trial_id, raw_material_id, quantity, lot_number, qty_per_small_unit, small_unit, phase)
                    VALUES (?, ?, ?, ?, ?, ?, ?)''',
                    (trial_id, ing['raw_id'], ing['total_needed'], ing['lot_number'], 
                     ing['qty_per_small'], ing['small_unit'], ing['phase']))

                if update_stock:
                    c.execute("SELECT unit, current_stock, name FROM raw_materials WHERE id = ?", (ing['raw_id'],))
                    mat = c.fetchone()
                    if mat:
                        converted = convert_quantity(ing['total_needed'], ing['small_unit'], mat['unit'])
                        c.execute('UPDATE raw_materials SET current_stock = current_stock - ?, updated_at = ? WHERE id = ?',
                                  (converted, now, ing['raw_id']))
                        log_stock_movement(
                            raw_material_id=ing['raw_id'],
                            movement_type='OUT',
                            quantity=converted,
                            unit=mat['unit'],
                            reference_type='TRIAL',
                            reference_id=trial_id,
                            reference_code=master_batch_id,
                            project_name=project_name,
                            notes=f"Used in trial {master_batch_id}",
                            lsa_number=ing.get('lot_number'),
                            conn=conn   # reuse koneksi transaksi utama
                        )

            # Catat riwayat sebelum commit agar tersimpan
            log_trial_history(
                trial_id=trial_id,
                action='CREATE',
                changed_by=session.get('name', 'System'),
                new_data={'project_name': project_name, 'master_batch_id': master_batch_id},
                notes='Trial dibuat pertama kali',
                conn=conn
            )

            conn.commit()

            flash(f'Trial "{master_batch_id}" berhasil dibuat!', 'success')
            return redirect(url_for('trial_detail', trial_id=trial_id))
        except Exception as e:
            conn.rollback()
            flash(f'Error: {str(e)}', 'danger')
            return redirect(url_for('create_trial'))
        finally:
            conn.close()

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, code, name, unit, current_stock FROM raw_materials ORDER BY name")
    raw_materials = c.fetchall()
    conn.close()
    return render_template('create_trial.html', raw_materials=raw_materials, today=date.today().isoformat())

@app.route('/trials/<int:trial_id>')
@login_required
def trial_detail(trial_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM trials WHERE id = ?", (trial_id,))
    trial = c.fetchone()

    if not trial:
        conn.close()
        flash('Trial tidak ditemukan.', 'danger')
        return redirect(url_for('trials_list'))

    c.execute("""
        SELECT ti.*, rm.code, rm.name, rm.unit 
        FROM trial_ingredients ti JOIN raw_materials rm ON ti.raw_material_id = rm.id
        WHERE ti.trial_id = ? ORDER BY ti.phase, ti.id
    """, (trial_id,))
    ingredients = c.fetchall()

    # Ambil riwayat perubahan
    c.execute("""
        SELECT * FROM trial_history 
        WHERE trial_id = ? 
        ORDER BY changed_at DESC
    """, (trial_id,))
    trial_history = c.fetchall()
    conn.close()

    return render_template('trial_detail.html', trial=trial, ingredients=ingredients, trial_history=trial_history)


@app.route('/trials/<int:trial_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_trial(trial_id):
    conn = get_db()
    c = conn.cursor()

    # Ambil data trial di awal (agar tersedia di POST dan GET)
    c.execute("SELECT * FROM trials WHERE id = ?", (trial_id,))
    trial = c.fetchone()

    if not trial:
        conn.close()
        flash('Trial tidak ditemukan.', 'danger')
        return redirect(url_for('trials_list'))

    if request.method == 'POST':
        batch_size_value = float(request.form.get('batch_size_value', 0) or 0)
        update_stock = request.form.get('update_stock') == 'on'
        now = datetime.now().isoformat()

        # Ambil data lama untuk audit log
        c.execute("SELECT * FROM trials WHERE id = ?", (trial_id,))
        old_trial = c.fetchone()

        try:
            # 1. Update header trial
            c.execute('''
                UPDATE trials SET
                    project_name = ?, objective = ?, batch_size_value = ?, batch_size_unit = ?,
                    scheme = ?, tahap = ?, procedure = ?, formulator = ?, updated_at = ?
                WHERE id = ?
            ''', (
                request.form.get('project_name', '').strip().upper(),
                request.form.get('objective', '').strip(),
                batch_size_value,
                request.form.get('batch_size_unit', 'tablet'),
                int(request.form.get('scheme', 1)),
                request.form.get('tahap', 'F1'),
                request.form.get('procedure', '').strip(),
                request.form.get('formulator', session.get('name', '')).strip().upper(),
                now, trial_id
            ))

            # 2. Ambil komposisi lama beserta satuan
            c.execute("SELECT raw_material_id, quantity, small_unit FROM trial_ingredients WHERE trial_id = ?", (trial_id,))
            old_ingredients = {}
            for row in c.fetchall():
                old_ingredients[row['raw_material_id']] = {
                    'quantity': row['quantity'],
                    'small_unit': row['small_unit'] or 'g'
                }

            # 3. Hapus komposisi lama
            c.execute("DELETE FROM trial_ingredients WHERE trial_id = ?", (trial_id,))

            # 4. Proses komposisi baru
            raw_ids = request.form.getlist('raw_material_id[]')
            qty_per_smalls = request.form.getlist('qty_per_small_unit[]')
            small_units = request.form.getlist('small_unit[]')
            phases = request.form.getlist('phase[]')
            lot_numbers = request.form.getlist('lot_number[]')

            # new_usage: {raw_material_id: {'quantity': total_in_small_unit, 'small_unit': unit}}
            new_usage = {}

            for i, raw_id in enumerate(raw_ids):
                if not raw_id or not qty_per_smalls[i]: continue
                qty_per_small = float(qty_per_smalls[i])
                small_unit = small_units[i] if i < len(small_units) else 'mg'
                phase = phases[i] if i < len(phases) else 'Fase Utama'
                lot_number = lot_numbers[i] if i < len(lot_numbers) else ''
                total_needed = round(qty_per_small * batch_size_value, 4)

                mat_id = int(raw_id)
                if mat_id not in new_usage:
                    new_usage[mat_id] = {'quantity': 0, 'small_unit': small_unit}
                new_usage[mat_id]['quantity'] += total_needed

                c.execute('''INSERT INTO trial_ingredients 
                    (trial_id, raw_material_id, quantity, lot_number, qty_per_small_unit, small_unit, phase)
                    VALUES (?, ?, ?, ?, ?, ?, ?)''',
                    (trial_id, mat_id, total_needed, lot_number, qty_per_small, small_unit, phase))

            # 5. Sesuaikan stok dengan konversi satuan yang akurat
            if update_stock:
                for mat_id, old_data in old_ingredients.items():
                    c.execute("SELECT unit, current_stock, name FROM raw_materials WHERE id = ?", (mat_id,))
                    mat = c.fetchone()
                    if mat:
                        # Konversi dari small_unit ke unit bahan baku
                        converted_old = convert_quantity(old_data['quantity'], old_data['small_unit'], mat['unit'])
                        c.execute('UPDATE raw_materials SET current_stock = current_stock + ?, updated_at = ? WHERE id = ?',
                                  (converted_old, now, mat_id))
                        log_stock_movement(mat_id, 'IN', converted_old, mat['unit'], 'ADJUSTMENT', trial_id,
                                           trial['master_batch_id'], trial['project_name'],
                                           f"Edit Trial - restored old usage", conn=conn)

                for mat_id, usage in new_usage.items():
                    c.execute("SELECT unit, current_stock, name FROM raw_materials WHERE id = ?", (mat_id,))
                    mat = c.fetchone()
                    if mat:
                        # Konversi akurat dari small_unit ke unit bahan baku
                        converted = convert_quantity(usage['quantity'], usage['small_unit'], mat['unit'])
                        if mat['current_stock'] < converted:
                            raise Exception(f"Stok {mat['name']} tidak mencukupi.")
                        c.execute('UPDATE raw_materials SET current_stock = current_stock - ?, updated_at = ? WHERE id = ?',
                                  (converted, now, mat_id))
                        log_stock_movement(mat_id, 'OUT', converted, mat['unit'], 'ADJUSTMENT', trial_id,
                                           trial['master_batch_id'], trial['project_name'],
                                           f"Edit Trial - new usage", conn=conn)

            # Catat riwayat perubahan sebelum commit agar tersimpan
            new_trial_data = {
                'project_name': request.form.get('project_name', '').strip(),
                'objective': request.form.get('objective', '').strip(),
                'batch_size_value': batch_size_value,
                'tahap': request.form.get('tahap', 'F1')
            }
            log_trial_history(
                trial_id=trial_id,
                action='UPDATE',
                changed_by=session.get('name', 'System'),
                old_data=dict(old_trial) if old_trial else None,
                new_data=new_trial_data,
                notes='Perubahan header & komposisi bahan',
                conn=conn
            )

            conn.commit()

            flash('Trial berhasil diperbarui termasuk komposisi bahan.', 'success')
            return redirect(url_for('trial_detail', trial_id=trial_id))

        except Exception as e:
            conn.rollback()
            flash(f'Gagal memperbarui trial: {str(e)}', 'danger')
            return redirect(url_for('edit_trial', trial_id=trial_id))
        finally:
            conn.close()

    # GET - Load ingredients untuk pre-fill form
    c.execute("""
        SELECT ti.*, rm.code, rm.name 
        FROM trial_ingredients ti
        JOIN raw_materials rm ON ti.raw_material_id = rm.id
        WHERE ti.trial_id = ?
        ORDER BY ti.phase, ti.id
    """, (trial_id,))
    current_ingredients = [dict(row) for row in c.fetchall()]
    conn.close()

    return render_template('edit_trial.html', trial=trial, current_ingredients=current_ingredients)


@app.route('/trials/<int:trial_id>/history')
@login_required
def trial_history(trial_id):
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM trials WHERE id = ?", (trial_id,))
    trial = c.fetchone()

    if not trial:
        conn.close()
        flash('Trial tidak ditemukan.', 'danger')
        return redirect(url_for('trials_list'))

    c.execute("""
        SELECT * FROM trial_history 
        WHERE trial_id = ? 
        ORDER BY changed_at DESC
    """, (trial_id,))
    history = c.fetchall()
    conn.close()

    return render_template('trial_history.html', trial=trial, history=history)


@app.route('/trials/<int:trial_id>/export_pdf')
@login_required
def export_trial_pdf(trial_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.units import cm, mm
    from reportlab.pdfgen import canvas
    import io
    from flask import make_response
    from datetime import datetime

    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM trials WHERE id = ?", (trial_id,))
    trial = c.fetchone()

    if not trial:
        conn.close()
        flash('Trial tidak ditemukan.', 'danger')
        return redirect(url_for('trials_list'))

    c.execute("""
        SELECT ti.*, rm.code, rm.name, rm.unit 
        FROM trial_ingredients ti 
        JOIN raw_materials rm ON ti.raw_material_id = rm.id
        WHERE ti.trial_id = ? ORDER BY ti.phase, ti.id
    """, (trial_id,))
    ingredients = c.fetchall()
    conn.close()

    buffer = io.BytesIO()

    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        page_num = canvas.getPageNumber()
        text = f"FormDX v1.0 | Halaman {page_num} | Dicetak: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        canvas.drawCentredString(A4[0]/2, 1*cm, text)
        canvas.restoreState()

    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()

    # Custom Styles
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18,
                                  textColor=colors.HexColor('#0d3b66'), spaceAfter=6, alignment=1)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10,
                                     textColor=colors.HexColor('#555555'), alignment=1, spaceAfter=15)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=11,
                                    textColor=colors.HexColor('#0d3b66'), spaceBefore=12, spaceAfter=6,
                                    borderPadding=4, backColor=colors.HexColor('#e8f4f8'))
    normal_style = ParagraphStyle('NormalText', parent=styles['Normal'], fontSize=9, leading=12)

    elements = []

    # === HEADER ===
    elements.append(Paragraph("CATATAN TRIAL FORMULASI", title_style))
    elements.append(Spacer(1, 0.4*cm))

    # === INFORMASI TRIAL ===
    elements.append(Paragraph("1. INFORMASI TRIAL", section_style))

    info_data = [
        ["No. Formula", trial['master_batch_id'] or trial['trial_code'], "Tanggal Trial", trial['trial_date']],
        ["Nama Produk", (trial['project_name'] or '').upper(), "Formulator", (trial['formulator'] or '-').upper()],
        ["Ukuran Batch", f"{trial['batch_size_value'] or 0} {trial['batch_size_unit']}", "Scheme / Tahap", f"{trial['scheme'] or '-'} / {trial['tahap'] or '-'}"]
    ]

    info_table = Table(info_data, colWidths=[3.5*cm, 6*cm, 3.5*cm, 6*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f4f8')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#f0f4f8')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*cm))

    # Objective
    if trial['objective']:
        elements.append(Paragraph("<b>Tujuan / Objective:</b>", normal_style))
        elements.append(Paragraph(trial['objective'], normal_style))
        elements.append(Spacer(1, 0.3*cm))

    # === KOMPOSISI BAHAN BAKU ===
    elements.append(Paragraph("2. KOMPOSISI BAHAN BAKU", section_style))

    if ingredients:
        ing_data = [["No", "Kode", "Nama Bahan Baku", "Qty per Satuan", "Satuan", "Total Kebutuhan"]]
        for idx, ing in enumerate(ingredients, 1):
            total = (ing['qty_per_small_unit'] or 0) * (trial['batch_size_value'] or 0)
            ing_data.append([
                str(idx),
                ing['code'],
                ing['name'][:35] + ("..." if len(ing['name']) > 35 else ""),
                f"{ing['qty_per_small_unit']:.3f}" if ing['qty_per_small_unit'] else "-",
                ing['small_unit'] or ing['unit'],
                f"{total:.3f} {ing['small_unit'] or ing['unit']}"
            ])

        ing_table = Table(ing_data, colWidths=[1*cm, 2.2*cm, 6*cm, 2.8*cm, 2*cm, 4.5*cm])
        ing_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d3b66')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (3, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#999999')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(ing_table)
    else:
        elements.append(Paragraph("<i>Tidak ada komposisi bahan.</i>", normal_style))

    elements.append(Spacer(1, 0.4*cm))

    # === PROSEDUR ===
    if trial['procedure']:
        elements.append(Paragraph("3. PROSEDUR PEMBUATAN", section_style))
        elements.append(Paragraph(trial['procedure'].replace('\n', '<br/>'), normal_style))

    # Build PDF
    doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)

    buffer.seek(0)
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    filename = f"Trial_{trial['master_batch_id'] or trial['trial_code']}.pdf"
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response


@app.route('/trials/<int:trial_id>/delete', methods=['POST'])
@login_required
def delete_trial(trial_id):
    conn = get_db()
    c = conn.cursor()
    try:
        # Ambil data trial
        c.execute("SELECT * FROM trials WHERE id = ?", (trial_id,))
        trial = c.fetchone()
        if not trial:
            flash('Trial tidak ditemukan.', 'danger')
            return redirect(url_for('trials_list'))

        # Ambil semua ingredients untuk mengembalikan stok (dengan satuan)
        c.execute("""
            SELECT ti.raw_material_id, ti.quantity, ti.small_unit, rm.unit, rm.name 
            FROM trial_ingredients ti
            JOIN raw_materials rm ON ti.raw_material_id = rm.id
            WHERE ti.trial_id = ?
        """, (trial_id,))
        ingredients = c.fetchall()

        now = datetime.now().isoformat()

        # Kembalikan stok dengan konversi satuan yang benar
        for ing in ingredients:
            converted_qty = convert_quantity(ing['quantity'], ing['small_unit'] or 'g', ing['unit'])
            c.execute('''
                UPDATE raw_materials 
                SET current_stock = current_stock + ?, updated_at = ?
                WHERE id = ?
            ''', (converted_qty, now, ing['raw_material_id']))

            c.execute('''
                INSERT INTO stock_movements 
                (raw_material_id, movement_type, quantity, unit, reference_type, reference_id, reference_code, project_name, notes, created_at)
                VALUES (?, 'IN', ?, ?, 'ADJUSTMENT', ?, ?, ?, ?, ?)
            ''', (
                ing['raw_material_id'],
                converted_qty,
                ing['unit'],
                trial_id,
                trial['master_batch_id'],
                trial['project_name'],
                f"Restored due to deletion of trial {trial['master_batch_id']}",
                now
            ))

        # Hapus trial (ingredients cascade delete)
        c.execute("DELETE FROM trials WHERE id = ?", (trial_id,))

        # Catat riwayat sebelum commit
        log_trial_history(
            trial_id=trial_id,
            action='DELETE',
            changed_by=session.get('name', 'System'),
            notes=f'Trial dihapus. Master Batch: {trial["master_batch_id"]}',
            conn=conn
        )

        conn.commit()

        flash(f'Trial "{trial["master_batch_id"]}" berhasil dihapus dan stok telah dikembalikan.', 'success')
        return redirect(url_for('trials_list'))

    except Exception as e:
        conn.rollback()
        flash(f'Gagal menghapus trial: {str(e)}', 'danger')
        return redirect(url_for('trial_detail', trial_id=trial_id))
    finally:
        conn.close()


# ==================== INBOUND STOCK ====================

@app.route('/raw_materials/<int:material_id>/inbound', methods=['GET', 'POST'])
@login_required
def inbound_stock(material_id):
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM raw_materials WHERE id = ?", (material_id,))
    material = c.fetchone()

    if not material:
        flash('Bahan baku tidak ditemukan.', 'danger')
        conn.close()
        return redirect(url_for('raw_materials'))

    if request.method == 'POST':
        try:
            quantity = float(request.form.get('quantity', 0))
            notes = request.form.get('notes', '').strip()
            reference_code = request.form.get('reference_code', '').strip()

            if quantity <= 0:
                flash('Jumlah harus lebih dari 0.', 'danger')
                return redirect(url_for('inbound_stock', material_id=material_id))

            now = datetime.now().isoformat()
            lpb_number = request.form.get('lpb_number', '').strip().upper()
            lsa_number = request.form.get('lsa_number', '').strip().upper()

            # Tambah stok
            c.execute('''
                UPDATE raw_materials 
                SET current_stock = current_stock + ?, updated_at = ?
                WHERE id = ?
            ''', (quantity, now, material_id))

            # Catat pergerakan stok dengan LPB & LSA
            c.execute('''
                INSERT INTO stock_movements 
                (raw_material_id, movement_type, quantity, unit, reference_type, reference_code, lpb_number, lsa_number, notes, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (material_id, 'IN', quantity, material['unit'], 'INBOUND', lpb_number, lpb_number, lsa_number, notes, now))

            conn.commit()
            flash(f'Stok {material["name"]} berhasil ditambahkan sebanyak {quantity} {material["unit"]}.', 'success')
            return redirect(url_for('stock_card', material_id=material_id))

        except Exception as e:
            conn.rollback()
            flash(f'Gagal menambahkan stok: {str(e)}', 'danger')
        finally:
            conn.close()

    conn.close()
    return render_template('inbound_stock.html', material=material)


@app.route('/raw_materials/<int:material_id>/adjustment', methods=['GET', 'POST'])
@login_required
def adjustment_stock(material_id):
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM raw_materials WHERE id = ?", (material_id,))
    material = c.fetchone()

    if not material:
        flash('Bahan baku tidak ditemukan.', 'danger')
        conn.close()
        return redirect(url_for('raw_materials'))

    if request.method == 'POST':
        try:
            quantity = float(request.form.get('quantity', 0))
            adjustment_type = request.form.get('adjustment_type', 'IN')  # IN atau OUT
            reason = request.form.get('reason', '').strip()
            notes = request.form.get('notes', '').strip()

            if quantity <= 0:
                flash('Jumlah harus lebih dari 0.', 'danger')
                return redirect(url_for('adjustment_stock', material_id=material_id))

            now = datetime.now().isoformat()

            # Update stok
            if adjustment_type == 'IN':
                c.execute('UPDATE raw_materials SET current_stock = current_stock + ?, updated_at = ? WHERE id = ?',
                          (quantity, now, material_id))
                movement_type = 'IN'
            else:
                c.execute('UPDATE raw_materials SET current_stock = current_stock - ?, updated_at = ? WHERE id = ?',
                          (quantity, now, material_id))
                movement_type = 'OUT'

            # Catat movement
            full_notes = f"Adjustment: {reason}. {notes}".strip()
            c.execute('''
                INSERT INTO stock_movements 
                (raw_material_id, movement_type, quantity, unit, reference_type, reference_code, notes, created_at)
                VALUES (?, ?, ?, ?, 'ADJUSTMENT', 'STOCK_ADJUSTMENT', ?, ?)
            ''', (material_id, movement_type, quantity, material['unit'], full_notes, now))

            conn.commit()
            flash(f'Stok {material["name"]} berhasil di-adjustment ({adjustment_type}) sebanyak {quantity} {material["unit"]}.', 'success')
            return redirect(url_for('stock_card', material_id=material_id))

        except Exception as e:
            conn.rollback()
            flash(f'Gagal melakukan adjustment: {str(e)}', 'danger')
        finally:
            conn.close()

    conn.close()
    return render_template('adjustment_stock.html', material=material)

# ==================== INISIALISASI DATABASE ====================

def init_db():
    conn = get_db()
    c = conn.cursor()

    # Tabel Bahan Baku
    c.execute('''
        CREATE TABLE IF NOT EXISTS raw_materials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            description TEXT,
            manufacturer TEXT,
            supplier TEXT,
            unit TEXT DEFAULT 'g',
            current_stock REAL DEFAULT 0,
            min_stock REAL DEFAULT 0,
            price_per_unit REAL DEFAULT 0,
            expiry_date TEXT,
            lpb_number TEXT,
            lsa_number TEXT,
            location TEXT DEFAULT 'R&D Coolroom',
            created_at TEXT,
            updated_at TEXT
        )
    ''')

    # Tabel Pergerakan Stok
    c.execute('''
        CREATE TABLE IF NOT EXISTS stock_movements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            raw_material_id INTEGER,
            movement_type TEXT NOT NULL,
            quantity REAL NOT NULL,
            unit TEXT,
            reference_type TEXT,
            reference_code TEXT,
            lpb_number TEXT,
            lsa_number TEXT,
            expiry_date TEXT,
            project_name TEXT,
            notes TEXT,
            created_at TEXT,
            FOREIGN KEY (raw_material_id) REFERENCES raw_materials (id)
        )
    ''')

    # Tabel Trial
    c.execute('''
        CREATE TABLE IF NOT EXISTS trials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            trial_code TEXT,
            master_batch_id TEXT,
            trial_date TEXT,
            project_name TEXT,
            formulator TEXT,
            status TEXT DEFAULT 'Draft',
            notes TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    ''')

    conn.commit()
    
# ==================== INISIALISASI DATABASE (WAJIB UNTUK PRODUCTION) ====================
init_db()


if __name__ == '__main__':
    # Hanya jalankan di development (lokal)
    with app.app_context():
        add_sample_data()   # Hanya dijalankan di lokal, bukan di Railway
    print("\n" + "="*65)
    print(" FormDX v1.0 - Clean Edition (Development Mode)")
    print(" http://127.0.0.1:5000")
    print(" Login: formulator / rd2026")
    print("="*65 + "\n")
    app.run(debug=True, host='0.0.0.0', port=5000)
